from requests_html import HTMLSession
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from openpyxl.utils.exceptions import InvalidFileException
import pandas as pd
import jmespath
import os
import json
import re


# Définition de la classe principale pour le scraping sur Liquipedia
class LiquipediaScraper:
    def __init__(self, base_url="https://liquipedia.net"):
        # Initialisation des paramètres pour le scraping, comme les tournois et les années d'intérêt
        parameters = {
            "Tiers": [
                "S-Tier_Tournaments",
                # "A-Tier_Tournaments",
                # "B-Tier_Tournaments",
                # "C-Tier_Tournaments",
            ],
            "Years": [
                "2024",
                "2023",
                # "2022",
                # "2021"
            ],
        }
        self.output_folder = "out/"
        if not os.path.exists(self.output_folder):
            os.makedirs(self.output_folder)
        self.base_url = base_url
        self.session = HTMLSession()
        self.excel_manager = ExcelManager()
        self.json_manager = JSONManager(self.excel_manager)
        self.image_downloader = ImageDownloader(self.session)
        self.team_scraper = TeamScraper(self.session, self.base_url, self.image_downloader, self.excel_manager, self.json_manager)
        self.tournament_scraper = TournamentScraper(
            self.session, self.base_url, self.excel_manager, self.json_manager, parameters["Tiers"], parameters["Years"]
        )

    def run(self):
        # Lance le processus de scraping pour les équipes et les tournois, puis sauvegarde les résultats dans un fichier Excel
        self.team_scraper.scrape_teams()
        self.tournament_scraper.scrape_tournaments()
        self.excel_manager.save()


class ExcelManager:
    def __init__(self, file_path="out/LiquipediaScraping.xlsx"):
        self.file_path = file_path
        self.workbook = None
        self.load_or_create_workbook()

    def load_or_create_workbook(self):
        if not os.path.exists(self.file_path):
            self.workbook = Workbook()
            print("Fichier Excel créé avec succès.")
        else:
            try:
                self.workbook = load_workbook(self.file_path)
                print("Fichier Excel chargé avec succès.")
            except InvalidFileException:
                print("Le fichier Excel est corrompu et ne peut pas être ouvert. Un nouveau fichier sera créé.")
                self.workbook = Workbook()
                print("Fichier Excel créé avec succès.")
        self.prepare_sheet()

    def prepare_sheet(self):
        if self.workbook.active.title == "Sheet":
            self.workbook.remove(self.workbook["Sheet"])
        if "Players" in self.workbook.sheetnames:
            self.workbook.remove(self.workbook["Players"])
        if "Tournaments" in self.workbook.sheetnames:
            self.workbook.remove(self.workbook["Tournaments"])
        if "Matches" in self.workbook.sheetnames:
            self.workbook.remove(self.workbook["Matches"])

        self.players_sheet = self.workbook.create_sheet(title="Players")
        self.players_sheet.append(["Flag", "Player Name", "Player Surname", "Role", "Team"])
        self.tournaments_sheet = self.workbook.create_sheet(title="Tournaments")
        self.matches_sheet = self.workbook.create_sheet(title="Matches")

    def save(self):
        self.workbook.save(self.file_path)
        print("Fichier Excel enregistré avec succès.")


class JSONManager:
    def __init__(self, excelmanager, file_path="out/tournament.json"):
        self.file_path = file_path
        self.excel_manager = excelmanager
        self.match_expression = "[*].matches[]"
        self.tournament_expression = "[*].{name : name, url : url, date:date, prize_pool:prize_pool,location:location,number_of_participants:number_of_participants,winner:winner,runner_up:runner_up}"

    def save(self, data):
        tournaments_data = self.save_tournaments_in_excel(data, self.excel_manager.tournaments_sheet)
        matches_data = self.save_matches_in_excel(data, self.excel_manager.matches_sheet)

        tournaments_data.to_json(self.file_path, orient="records")
        matches_data.to_json(self.file_path, orient="records")
        print("Fichier JSON enregistré avec succès.")

        

    def save_tournaments_in_excel(self, data, sheet):
        # Ajout des tournaments
        tournaments = jmespath.search(self.tournament_expression, data)
        tournaments_data = pd.DataFrame(tournaments)

        if "url" in tournaments_data.columns:
            tournaments_data["url"] = "https://liquipedia.net" + tournaments_data["url"]
            
        sheet.append(tournaments_data.columns.tolist())
        for row in tournaments_data.iterrows():
            sheet.append(row[1].tolist())
        return tournaments_data

    def save_matches_in_excel(self, data, sheet):
        # Ajout des matches
        matches = jmespath.search(self.match_expression, data)
        matches_data = pd.DataFrame(matches)

        if "siegegg" in matches_data.columns:
            matches_data["siegegg"] = "https://siege.gg/matches/" + matches_data["siegegg"]
        if "r6esports" in matches_data.columns:
            matches_data["r6esports"] = "https://www.ubisoft.com/en-us/esports/rainbow-six/siege/match/" + matches_data["r6esports"]

        sheet.append(matches_data.columns.tolist())
        for row in matches_data.iterrows():
            sheet.append(row[1].tolist())
        return matches_data

    @staticmethod
    def flatten_json(json_input, parent_key="", sep="-"):
        items = []
        if isinstance(json_input, dict):
            for k, v in json_input.items():
                new_key = f"{parent_key}{sep}{k}" if parent_key else k
                if isinstance(v, dict):
                    items.extend(JSONManager.flatten_json(v, new_key, sep=sep).items())
                elif isinstance(v, list):
                    for item in v:
                        # Traitement des éléments de la liste sans ajouter d'indice
                        items.extend(JSONManager.flatten_json(item, new_key, sep=sep).items())
                else:
                    items.append((new_key, v))
        elif isinstance(json_input, list):
            for item in json_input:
                # Même traitement pour les listes à la racine (bien que peu commun)
                items.extend(JSONManager.flatten_json(item, parent_key, sep=sep).items())
        return dict(items)


class ImageDownloader:
    def __init__(self, session, media_folder="media/"):
        self.media_folder = media_folder
        self.session = session
        if not os.path.exists(self.media_folder):
            os.makedirs(self.media_folder)
        if not os.path.exists(self.media_folder + "TeamLogo"):
            os.makedirs(self.media_folder + "TeamLogo")
        if not os.path.exists(self.media_folder + "PlayerFlag"):
            os.makedirs(self.media_folder + "PlayerFlag")

    @staticmethod
    def get_image_name_from_url(url):
        return url.split("/")[-1]

    def download_image(self, folder, image_url):
        r = self.session.get(image_url)
        image_name = self.get_image_name_from_url(image_url)
        with open(self.media_folder + folder + image_name, "wb") as f:
            f.write(r.content)
        f.close()


class TeamScraper:
    def __init__(self, session, base_url, image_downloader, excel_manager, json_manager):
        self.session = session
        self.base_url = base_url
        self.image_downloader = image_downloader
        self.excel_manager = excel_manager
        self.json_manager = json_manager

    def scrape_teams(self, teams_url="https://liquipedia.net/rainbowsix/Portal:Teams"):
        r = self.session.get(teams_url)
        # Recupère la liste des régions
        region_html_list = r.html.find("div.tabs-static li")
        region_list = []
        for region_html in region_html_list:
            region = region_html.find("a", first=True).full_text
            if region != "Overview" and region != "Players":
                region_list.append(region)

        # Recupère la liste des équipes par région
        for region in region_list:
            url_region = teams_url + "/" + region
            r = self.session.get(url_region)
            team_html_list = r.html.find("div.template-box table.wikitable")

            for team_html in team_html_list:
                team_logo = team_html.find("span.team-template-image-icon img", first=True).attrs["src"]
                self.image_downloader.download_image("TeamLogo/", self.base_url + team_logo)

                team_name = team_html.find("span.team-template-text", first=True).full_text
                player_html_list = team_html.find("tr")[2:]
                for player_html in player_html_list:
                    player_flag = player_html.find("td img")[0].attrs["src"]
                    player_surname = player_html.find("td")[0].full_text
                    player_name = player_html.find("td")[1].full_text
                    player_role = player_html.find("td")[2].full_text

                    # download playerflag image si elle n'existe pas
                    if not os.path.exists(self.image_downloader.media_folder + self.image_downloader.get_image_name_from_url(player_flag)):
                        self.image_downloader.download_image("PlayerFlag/", self.base_url + player_flag)

                    # Insertion dans le fichier excel
                    self.excel_manager.players_sheet.append(["", player_name, player_surname, player_role, team_name])
                    # insertion de l'image dans le fichier excel
                    flag = Image(self.image_downloader.media_folder + "PlayerFlag/" + self.image_downloader.get_image_name_from_url(player_flag))
                    self.excel_manager.players_sheet.add_image(
                        flag,
                        get_column_letter(1) + str(self.excel_manager.players_sheet.max_row),
                    )
                print(f"Les joueurs de l'équipe {team_name} ont été extaits avec succès.")


class TournamentScraper:
    def __init__(
        self,
        session,
        base_url,
        excel_manager,
        json_manager,
        list_tiers=["S-Tier_Tournaments"],
        years=["2024"],
    ):
        self.session = session
        self.base_url = base_url
        self.r6_base_url = base_url + "/rainbowsix/"
        self.list_tiers = list_tiers
        self.years = years
        self.tournament_result = []
        self.json_manager = json_manager
        self.excelmanager = excel_manager

    def scrape_tournaments(self):
        if len(self.list_tiers) == 0 or self.list_tiers == None:
            print("Aucun tournoi n'a été trouvé.")
            return

        html_tournament_headers = [
            "G & S",
            "Tournament",
            "Date",
            "Prize\xa0Pool",
            "Location",
            "P#",
            "Winner",
            "Runner-up",
        ]

        for tiers in self.list_tiers:
            url_tier = self.r6_base_url + tiers
            r = self.session.get(url_tier)

            html_tournament_headers = []
            headers_html = r.html.find("div.mw-parser-output div.gridTable.tournamentCard", first=True).find("div.gridHeader div.gridCell")
            for header in headers_html:
                html_tournament_headers.append(header.full_text)

            tournaments_html_list = r.html.find("div.mw-parser-output div.gridTable.tournamentCard div.gridRow")
            for tournament_html in tournaments_html_list:
                # create tournament object with name, date, prize_pool, location, number_of_participants, winner, runner_up
                tournament = {
                    "name": tournament_html.find("div.gridCell.Tournament a")[-1].full_text,
                    "url": tournament_html.find("div.gridCell.Tournament a")[-1].attrs["href"],
                    "date": tournament_html.find("div.gridCell.Date", first=True).full_text,
                    "prize_pool": tournament_html.find("div.gridCell.Prize", first=True).full_text,
                    "location": tournament_html.find("div.gridCell.Location", first=True).full_text,
                    "number_of_participants": tournament_html.find("div.gridCell.PlayerNumber", first=True).full_text,
                }
                try:
                    tournament["winner"] = tournament_html.find("div.gridCell.FirstPlace a")[-1].full_text
                    tournament["runner_up"] = tournament_html.find("div.gridCell.SecondPlace a")[-1].full_text
                except:
                    tournament["winner"] = ""
                    tournament["runner_up"] = ""

                for year in self.years:
                    if year in tournament["date"]:
                        if tournament["winner"] != "":
                            match_scraper = MatchScraper(self.session, self.base_url, tournament, self.json_manager)
                            result = match_scraper.extract_matches()
                            # apprend avec le nom du tournoi seulement
                            tournament["matches"] = result
                            self.tournament_result.append(tournament)
                            print(f"Le tournoi {tournament['name']} a été ajouté.")
                        else:
                            print(f"Le tournoi {tournament['name']} est annulé ou en cours.")

        self.json_manager.save(self.tournament_result)


class MatchScraper:
    def __init__(
        self,
        session,
        base_url,
        tournament,
        json_manager,
        api_url="https://liquipedia.net/rainbowsix/api.php?action=query&format=json&prop=revisions&rvprop=content&titles=",
    ):
        self.session = session
        self.base_url = base_url
        self.tournament = tournament
        self.api_url = api_url
        self.data = self.session.get(self.api_url + self.tournament["url"].replace("/rainbowsix/", "")).json()
        self.match_section = ""
        self.pattern = r"\{\{#section:[^}]+\}\}"
        self.pattern_result = ""
        self.match_result = []
        self.jsonconvert = MatchToJsonConverter()
        self.json_manager = json_manager

    def extract_match_section(self, match_block):
        print("MAtchSection")
        match_block = match_block.replace("\\n", "").replace("\\t", "").replace("\\", "").replace("{{", "").replace("}}", "")
        parts = match_block.split("|")
        if len(parts) > 1:
            self.match_section = parts[1]

    def extract_matches(self):
        data_str = str(self.data["query"]["pages"]).replace("\\n", "").replace("\\t", "")
        if "Match|" in data_str or "BracketMatchSummary|" in data_str:
            self.extract_json_matches(data_str)
        else:
            tournament = str(self.tournament["url"].replace("/rainbowsix/", ""))
            self.pattern_result = re.findall(self.pattern, data_str)
            for section in self.pattern_result:
                section = section.split("section:")[-1].replace(tournament, "").split("|")[0]
                self.data = self.session.get(self.api_url + self.tournament["url"].replace("/rainbowsix/", "") + section).json()
                data_str = str(self.data["query"]["pages"]).replace("\\n", "").replace("\\t", "")
                self.extract_json_matches(data_str)

        # dans tous les objets ayant une clé date, replace
        for match in self.match_result:
            try:
                match["date"] = match["date"].replace(",", "")
            except:
                pass
        return self.match_result

    def extract_json_matches(self, data_str):
        matches = []
        pos = 0
        while pos < len(data_str):
            start = data_str.find("{{Match|", pos)
            if start == -1:
                start = data_str.find("{{BracketMatchSummary|", pos)
                if start == -1:
                    break

            count = 0
            for i in range(start, len(data_str)):
                if data_str[i : i + 2] == "{{":
                    count += 1
                elif data_str[i : i + 2] == "}}":
                    count -= 1
                if count == 0:
                    match_block = data_str[start : i + 2]
                    if "MatchSection" in match_block:
                        self.extract_match_section(match_block)
                        pos = i + 2
                        break
                    matches.append(match_block)
                    pos = i + 2
                    break
            else:
                raise ValueError("Match block not properly closed.")

        print(f"{len(matches)} matchs ont été trouvés.")
        if len(matches) == 0:
            print("\n\nAucun match n'a été trouvé.")
            return

        for match in matches:
            json_output = self.jsonconvert.extract_objects(match, self.match_section)
            json_output = self.json_manager.flatten_json(json_output)
            self.match_result.append(json_output)
        return self.match_result


class MatchToJsonConverter:
    @staticmethod
    def text_to_json(text, match_section=""):
        obj = MatchToJsonConverter.parse_object(text)
        if match_section:
            obj["MatchSection"] = match_section
        return obj

    @staticmethod
    def parse_object(s):
        temp_key = ""
        objects = {}
        for i in range(len(s)):
            if "=" not in s[i]:
                try:
                    objects[temp_key] += "|" + s[i]
                except:
                    continue
                continue
            key, value = s[i].split("=", 1)
            if "[" in value and "|" in value:
                value = value.replace("[", "").replace("]", "").split("|")
                value = value[1:]
                if len(value) == 2 and "opponent" in key:
                    value = value[:1]
                if len(value) > 1:
                    value = [MatchToJsonConverter.parse_object(value)]
                else:
                    value = value[0]
            objects[key] = value
            temp_key = key
        return objects

    @staticmethod
    def extract_objects(s, match_section=""):
        s = s.replace("\\n", "").replace("\\t", "").replace("\\", "").replace("{{", "[").replace("}}", "]")
        s = s.split("|")[1:]  # Remove first element which is 'Match'
        count_brackets = 0
        new_string = []
        for part in s:
            if count_brackets > 0:
                new_string[-1] += "|" + part
            else:
                new_string.append(part)
            count_brackets += part.count("[") - part.count("]")
        return MatchToJsonConverter.text_to_json(new_string, match_section)


scraper = LiquipediaScraper()
scraper.run()
